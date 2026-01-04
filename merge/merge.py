# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
import xlrd
import os
import glob
import re
from datetime import datetime

script_dir = os.path.dirname(os.path.abspath(__file__))

print("=" * 60)
print("Excel Merge Tool")
print("=" * 60)

excel_files = []
for pattern in ['*.xls', '*.xlsx', '*.xlsm']:
    for f in glob.glob(os.path.join(script_dir, pattern)):
        filename = os.path.basename(f)
        if not filename.startswith('huizong'):
            excel_files.append(f)

# Windows natural sort: 1-1, 1-2, ..., 1-9, 1-10, 1-11
def natural_key(path):
    name = os.path.basename(path)
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', name)]

excel_files.sort(key=natural_key)

if not excel_files:
    print("\nNo Excel files found!")
    input("\nPress Enter...")
    exit()

print(f"\nFound {len(excel_files)} files")

wb_out = Workbook()
ws_out = wb_out.active

ws_out.cell(row=1, column=1, value='日期')
ws_out.cell(row=1, column=2, value='银行')
ws_out.cell(row=1, column=3, value='摘要')
ws_out.cell(row=1, column=4, value='借方')
ws_out.cell(row=1, column=5, value='贷方')
ws_out.cell(row=1, column=6, value='余额')

out_row = 2
total = 0
skipped_files = []

def open_excel_file(file_path):
    """Open Excel file and return (book, file_type) tuple"""
    if file_path.lower().endswith('.xls'):
        return xlrd.open_workbook(file_path), 'xls'
    else:  # .xlsx, .xlsm
        return load_workbook(file_path, data_only=True), 'xlsx'

def get_value(sheet, r, c, book, file_type):
    """Get cell value from sheet, handling both xls and xlsx formats"""
    if file_type == 'xls':
        cell = sheet.cell(r, c)
        if cell.ctype == 3:  # Date type
            try:
                dt = xlrd.xldate_as_datetime(cell.value, book.datemode)
                return dt
            except:
                return cell.value
        elif cell.ctype == 0:  # Empty
            return ''
        else:
            return cell.value
    else:  # xlsx
        cell = sheet.cell(r + 1, c + 1)  # openpyxl uses 1-based indexing
        if cell.value is None:
            return ''
        elif isinstance(cell.value, datetime):
            return cell.value
        else:
            return cell.value

# Fixed column positions: A=日期(0) B=银行(1) C=摘要(2) D=借方(3) E=贷方(4) F=余额(5)
COL_DATE = 0
COL_BANK = 1
COL_SUMMARY = 2
COL_DEBIT = 3
COL_CREDIT = 4
COL_BALANCE = 5

for file_path in excel_files:
    filename = os.path.basename(file_path)
    print(f"\n[{filename}]")

    file_total = 0
    try:
        book, file_type = open_excel_file(file_path)

        # Get sheet list based on file type
        if file_type == 'xls':
            sheets = [(book.sheet_by_index(i), book.sheet_by_index(i).name) for i in range(book.nsheets)]
            get_nrows = lambda sheet: sheet.nrows
            get_ncols = lambda sheet: sheet.ncols
        else:  # xlsx
            sheets = [(book[name], name) for name in book.sheetnames]
            get_nrows = lambda sheet: sheet.max_row
            get_ncols = lambda sheet: sheet.max_column

        # Loop through all sheets
        for sheet_idx, (sheet, sheet_name) in enumerate(sheets):
            if sheet_idx > 0:
                print(f"  Sheet: {sheet_name}")

            count = 0
            # Process all rows, skip rows where A or F is empty
            for r in range(get_nrows(sheet)):
                # Check if column A and F both have values
                if get_ncols(sheet) <= COL_BALANCE:
                    continue

                date_val = get_value(sheet, r, COL_DATE, book, file_type)
                bal_val = get_value(sheet, r, COL_BALANCE, book, file_type)

                # Only include rows where A and F both have values
                if date_val != '' and bal_val != '':
                    # Filter: only 2025-12 and later
                    year = date_val.year if hasattr(date_val, 'year') else 0
                    month = date_val.month if hasattr(date_val, 'month') else 0
                    # Skip only if year/month extraction failed OR date is before 2025-12
                    if (year > 0 and month > 0) and (year > 2025 or (year == 2025 and month == 12)):
                        ws_out.cell(row=out_row, column=1, value=date_val)
                        ws_out.cell(row=out_row, column=2, value=get_value(sheet, r, COL_BANK, book, file_type))
                        ws_out.cell(row=out_row, column=3, value=get_value(sheet, r, COL_SUMMARY, book, file_type))
                        ws_out.cell(row=out_row, column=4, value=get_value(sheet, r, COL_DEBIT, book, file_type))
                        ws_out.cell(row=out_row, column=5, value=get_value(sheet, r, COL_CREDIT, book, file_type))
                        ws_out.cell(row=out_row, column=6, value=bal_val)
                        out_row += 1
                        count += 1

            if count > 0:
                print(f"  Copied: {count} (Sheet: {sheet_name})")
            file_total += count
            total += count

        # Track files with no data
        if file_total == 0:
            skipped_files.append(filename)
            print(f"  No data copied")

    except Exception as e:
        print(f"  Error: {e}")
        skipped_files.append(filename)

print("\n" + "-" * 60)
print(f"Files processed: {len(excel_files)}")
print(f"Total rows: {total}")

if skipped_files:
    print(f"\nSkipped files ({len(skipped_files)}):")
    for f in skipped_files:
        print(f"  - {f}")

output_file = os.path.join(script_dir, 'huizong.xlsx')
try:
    wb_out.save(output_file)
    print(f"\nDone! Output: huizong.xlsx")
    os.system(f'explorer /select,"{output_file}"')
except PermissionError:
    print("\nError: Close huizong.xlsx first!")

input("\nPress Enter...")
